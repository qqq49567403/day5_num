# handle_excel.py
# 读取excel文件，进行key-value进行拼接

from openpyxl import load_workbook



class HandleExcel:
    def __init__(self, excel_path, sheet):
        self.wb = load_workbook(excel_path)
        self.sh = self.wb[sheet]

    # 获取excel所有行数据
    def __get_all(self):
        self.all_row = list(self.sh.rows)

    # 获取第一行数据作为title
    def _get_titile(self):
        # 获取所有行数据
        self.__get_all()
        # 获取第一行数据作为key值
        title = []
        for item in self.all_row[0]:
            title.append(item.value)
        return title

    # 获取所有行数据，与title拼接，以key-value的形式
    def get_all_data(self):
        # 获取title信息
        title = self._get_titile()
        # 测试数据集
        case_data = []
        # 遍历测试数据与title进行拼接
        for item in self.all_row[1:]:
            value = []
            for cell in item:
                value.append(cell.value)
            case = dict(zip(title, value))
            case_data.append(case)
        return case_data

    # 添加数据到excel表中
    def write_data(self, row, col, value):
        self.sh.cell(row, col).value = value

    # 保存excel数据
    def save_data(self, filename):
        self.wb.save(filename)


if __name__ == '__main__':
    import os
    from common.handle_path import testdata_dir

    excel_path = os.path.join(testdata_dir, 'api_cases.xlsx')
    he = HandleExcel(excel_path, '注册')
    cases = he.get_all_data()
    print(cases)
